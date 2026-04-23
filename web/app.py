import os, sys, io, json
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from flask import Flask, render_template_string, request, send_file, Response
from markupsafe import Markup
import functools
from datetime import date, timedelta
from db.database import (
    get_daily_report, get_mixer_shifts, get_pallets, get_tasks,
    get_recipe, save_task, update_recipe, init_db,
    get_all_users, revoke_user, get_codes, update_code,
    get_transit_pallets, get_warehouse_summary, get_warehouse_total,
    get_warehouse_pallets, get_lacquer_records, get_full_export,
    get_active_tasks, complete_task_pallet, save_pallet
)

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "lvt-2026")
WEB_PASSWORD = os.getenv("WEB_PASSWORD", "admin2026")
WEB_LOGIN = os.getenv("WEB_LOGIN", "admin")

def require_auth(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        # Check cookie
        if request.cookies.get("auth") == f"{WEB_LOGIN}:{WEB_PASSWORD}":
            return f(*args, **kwargs)
        # Check form submission
        if request.method == "POST" and request.form.get("_login"):
            if (request.form.get("_login") == WEB_LOGIN and
                    request.form.get("_password") == WEB_PASSWORD):
                resp = f(*args, **kwargs)
                from flask import make_response
                r = make_response(resp)
                r.set_cookie("auth", f"{WEB_LOGIN}:{WEB_PASSWORD}",
                             max_age=86400*30, httponly=True)
                return r
        # Show login form
        return _login_page()
    return decorated

def _login_page(error=""):
    from flask import Response as R
    html = f"""<!DOCTYPE html><html lang="ru"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>LVT - Вход</title>
<style>*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:system-ui,sans-serif;background:#f5f5f0;display:flex;align-items:center;justify-content:center;min-height:100vh}}
.box{{background:#fff;border:1px solid #e0ddd5;border-radius:12px;padding:40px;width:320px;text-align:center}}
h1{{font-size:18px;color:#185FA5;margin-bottom:6px}}p{{font-size:13px;color:#6b6b66;margin-bottom:20px}}
input{{width:100%;padding:10px;border:1px solid #e0ddd5;border-radius:8px;font-size:14px;margin-bottom:10px}}
button{{width:100%;padding:10px;background:#185FA5;color:#fff;border:none;border-radius:8px;font-size:14px;cursor:pointer}}
.err{{color:#A32D2D;font-size:13px;margin-bottom:10px}}</style></head>
<body><div class="box">
<h1>LVT Production</h1><p>Panel</p>
{"<div class=err>Неверный логин или пароль</div>" if error else ""}
<form method="post">
<input name="_login" placeholder="Логин" required>
<input name="_password" type="password" placeholder="Пароль" required>
<button type="submit">Войти</button>
</form></div></body></html>"""
    return R(html, 200, {{"Content-Type": "text/html; charset=utf-8"}})

NAV = [
    ("/", "Отчёт"),
    ("/week", "Неделя"),
    ("/transit", "🏭 Транзит"),
    ("/warehouse", "🏪 Склад"),
    ("/mixer", "Миксер"),
    ("/extruder", "Экструзия"),
    ("/lacquer", "Лак"),
    ("/tasks", "Задание"),
    ("/recipe", "Рецептура"),
    ("/users", "👥 Пользователи"),
    ("/export", "📥 Экспорт"),
]

CSS = """
<style>
*{box-sizing:border-box;margin:0;padding:0}
:root{--bg:#f5f5f0;--sur:#fff;--bor:#e0ddd5;--txt:#1a1a18;--mut:#6b6b66;
      --acc:#185FA5;--grn:#3B6D11;--grn-bg:#EAF3DE;--amb:#854F0B;--amb-bg:#FAEEDA;
      --red:#A32D2D;--red-bg:#FCEBEB;--transit:#7B3F00;--transit-bg:#FFF3E0}
body{font-family:system-ui,-apple-system,sans-serif;background:var(--bg);color:var(--txt);font-size:14px;line-height:1.5}
.top{background:var(--sur);border-bottom:1px solid var(--bor);padding:10px 20px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px}
.top h1{font-size:15px;font-weight:500;color:var(--acc)}
.nav{display:flex;gap:4px;flex-wrap:wrap}
.nav a{padding:5px 11px;border-radius:8px;text-decoration:none;font-size:12px;color:var(--mut);border:1px solid var(--bor);white-space:nowrap}
.nav a.active,.nav a:hover{background:var(--acc);color:#fff;border-color:var(--acc)}
.con{max-width:1100px;margin:0 auto;padding:20px 14px}
.g4{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:10px;margin-bottom:18px}
.card{background:var(--sur);border:1px solid var(--bor);border-radius:10px;padding:14px 16px;margin-bottom:12px}
.card h2{font-size:13px;font-weight:500;margin-bottom:10px;padding-bottom:8px;border-bottom:1px solid var(--bor)}
.met{background:var(--sur);border:1px solid var(--bor);border-radius:10px;padding:12px 14px}
.met .lb{font-size:11px;color:var(--mut);margin-bottom:3px}
.met .vl{font-size:26px;font-weight:500}
.met .sb{font-size:11px;color:var(--mut);margin-top:2px}
.met.acc .vl{color:var(--acc)}.met.red .vl{color:var(--red)}.met.grn .vl{color:var(--grn)}
.met.transit{border-color:#FF9800}.met.transit .vl{color:var(--transit)}
table{width:100%;border-collapse:collapse;font-size:12px}
th{text-align:left;padding:7px 8px;font-size:10px;font-weight:500;color:var(--mut);border-bottom:1px solid var(--bor);text-transform:uppercase;letter-spacing:.3px}
td{padding:7px 8px;border-bottom:1px solid var(--bor)}
tr:last-child td{border-bottom:none}tr:hover td{background:var(--bg)}
.badge{display:inline-block;padding:2px 7px;border-radius:5px;font-size:10px;font-weight:500}
.ok{background:var(--grn-bg);color:var(--grn)}.warn{background:var(--amb-bg);color:var(--amb)}
.err{background:var(--red-bg);color:var(--red)}.day{background:var(--amb-bg);color:var(--amb)}
.night{background:#E6F1FB;color:var(--acc)}
.transit-badge{background:var(--transit-bg);color:var(--transit)}
.warehouse-badge{background:var(--grn-bg);color:var(--grn)}
.prog-wrap{margin-bottom:10px}
.prog-lbl{display:flex;justify-content:space-between;font-size:11px;margin-bottom:3px}
.prog-bar{height:7px;background:var(--bg);border-radius:4px;overflow:hidden}
.prog-fill{height:100%;border-radius:4px}
.two{display:grid;grid-template-columns:1fr 1fr;gap:12px}
@media(max-width:640px){.two{grid-template-columns:1fr}}
.btn{padding:7px 14px;border:1px solid var(--bor);border-radius:8px;background:var(--sur);font-size:13px;cursor:pointer;text-decoration:none;display:inline-block;color:var(--txt)}
.btn:hover{background:var(--bg)}.btn-p{background:var(--acc);color:#fff;border-color:var(--acc)}.btn-p:hover{opacity:.9}
.btn-grn{background:var(--grn);color:#fff;border-color:var(--grn)}
input,select,textarea{padding:7px 10px;border:1px solid var(--bor);border-radius:8px;font-size:13px;background:var(--sur);color:var(--txt);width:100%}
.fr{display:flex;flex-direction:column;gap:4px;margin-bottom:10px}
.fr label{font-size:12px;color:var(--mut)}
.dbar{display:flex;gap:8px;align-items:center;margin-bottom:14px;flex-wrap:wrap}
.transit-alert{background:var(--transit-bg);border:1px solid #FF9800;border-radius:10px;padding:12px 16px;margin-bottom:16px;color:var(--transit);font-size:13px}
.ai-box{background:#f0f4ff;border:1px solid #c0cef0;border-radius:10px;padding:14px 16px;margin-top:12px}
.ai-resp{background:var(--sur);border:1px solid var(--bor);border-radius:8px;padding:12px;margin-top:10px;font-size:13px;line-height:1.7;white-space:pre-wrap;min-height:60px}
</style>
"""



def nav_html(active):
    links = "".join(f'<a href="{u}" class="{"active" if u==active else ""}">{l}</a>' for u,l in NAV)

    transit = get_transit_pallets()
    alert = f'<span style="background:#FF9800;color:#fff;padding:3px 8px;border-radius:6px;font-size:11px;margin-left:8px">⚠️ Транзит: {len(transit)}</span>' if transit else ""
    return f'<div class="top"><h1>ЛВТ Производство{alert}</h1><nav class="nav">{links}</nav></div>'

def badge(pct):
    cls = "ok" if pct<=2 else ("warn" if pct<=5 else "err")
    return f'<span class="badge {cls}">{pct}%</span>'

def prog_color(pct):
    return "#639922" if pct>=80 else ("#EF9F27" if pct>=50 else "#E24B4A")

def render(active, content):
    return render_template_string(CSS + nav_html(active) + '<div class="con">'+content+'</div>', content=Markup(content))


# ── ОТЧЁТ ─────────────────────────────────────────────────────────────────────

@app.route("/")
@require_auth
def report():
    ds = request.args.get("date", date.today().strftime("%d.%m.%Y"))
    r = get_daily_report(ds)
    pct = round(r["total_defect"]/r["total_qty"]*100,1) if r["total_qty"] else 0
    transit = get_transit_pallets()

    decor_rows = "".join(f'<tr><td><b>{d["decor"]}</b></td><td>{d["qty"]}</td>'
                         f'<td>{d["defect"]} {badge(round(d["defect"]/d["qty"]*100,1) if d["qty"] else 0)}</td></tr>'
                         for d in r["by_decor"])
    recipe_rows = "".join(f'<tr><td>{c["component"]}</td><td>{c["kg_per_batch"]} кг</td>'
                          f'<td><b>{c["kg_per_batch"]*r["batches"]:.1f} кг</b></td></tr>'
                          for c in r["recipe"])
    remark_rows = "".join(f'<tr><td>{rm["remark_type"]}</td><td>{rm["qty"]}</td><td>{rm["reason"]}</td></tr>'
                          for rm in r["remarks"])

    transit_block = ""
    if transit:
        transit_block = f'<div class="transit-alert">🏭 В транзитной зоне: <b>{len(transit)} паллет</b> ожидают обработки лаком</div>'

    content = f"""
    {transit_block}
    <div class="dbar"><span>Дата:</span>
      <form method="get" style="display:flex;gap:6px">
        <input name="date" value="{ds}" style="width:120px">
        <button type="submit" class="btn btn-p">Показать</button>
      </form>
    </div>
    <div class="g4">
      <div class="met acc"><div class="lb">Выпуск листов</div><div class="vl">{r["total_qty"]}</div><div class="sb">{ds}</div></div>
      <div class="met {"red" if pct>5 else "grn"}"><div class="lb">Брак</div><div class="vl">{r["total_defect"]}</div><div class="sb">{pct}% от выпуска</div></div>
      <div class="met"><div class="lb">Замесов</div><div class="vl">{r["batches"]}</div><div class="sb">{r["mixer_detail"]}</div></div>
      <div class="met transit"><div class="lb">В транзите</div><div class="vl">{len(transit)}</div><div class="sb">паллет ждут лак</div></div>
      <div class="met grn"><div class="lb">Лак обработано</div><div class="vl">{r["lac_pallets"]}</div><div class="sb">{r["lac_qty"]} листов</div></div>
      <div class="met"><div class="lb">Сырьё</div><div class="vl">{r["total_kg"]/1000:.1f}</div><div class="sb">тонн ({r["total_kg"]:.0f} кг)</div></div>
    </div>
    <div class="two">
      <div class="card"><h2>Выпуск по декорам</h2>
        {"<table><thead><tr><th>Декор</th><th>Листов</th><th>Брак</th></tr></thead><tbody>"+decor_rows+"</tbody></table>" if decor_rows else "<p style='color:var(--mut);font-size:13px'>Нет данных</p>"}
      </div>
      <div class="card"><h2>Расход сырья ({r["batches"]} замесов)</h2>
        {"<table><thead><tr><th>Компонент</th><th>На замес</th><th>Итого</th></tr></thead><tbody>"+recipe_rows+"</tbody></table>" if r["batches"] else "<p style='color:var(--mut);font-size:13px'>Нет замесов</p>"}
      </div>
    </div>
    <div class="card"><h2>Замечания по смене</h2>
      {"<table><thead><tr><th>Тип</th><th>Кол-во</th><th>Причина</th></tr></thead><tbody>"+remark_rows+"</tbody></table>" if remark_rows else "<p style='color:var(--mut);font-size:13px'>Нет замечаний</p>"}
    </div>"""
    return render("/", content)


# ── ТРАНЗИТ ───────────────────────────────────────────────────────────────────

@app.route("/transit")
@require_auth
def transit_page():
    pallets = get_transit_pallets()
    rows = ""
    for p in pallets:
        rows += f"""<tr>
          <td><b>ID{p["id"]}</b></td>
          <td>{p["date"]}</td>
          <td><span class="badge {"day" if p["shift"]=="день" else "night"}">{p["shift"].capitalize()}</span></td>
          <td><b>{p["decor"]}</b></td>
          <td>#{p["pallet_num"]}</td>
          <td><b>{p["qty"]}</b></td>
          <td>{"<span class='badge err'>"+str(p["defect"])+"</span>" if p["defect"] else "0"}</td>
          <td style="color:var(--mut)">{p["operator"]}</td>
          <td><span class="badge transit-badge">Ожидает лак</span></td>
        </tr>"""

    content = f"""
    <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:14px;flex-wrap:wrap;gap:8px">
      <h2 style="font-size:16px">🏭 Транзитная зона — {len(pallets)} паллет</h2>
      <a href="/warehouse" class="btn btn-grn">🏪 Перейти на склад</a>
    </div>
    {"<div class='transit-alert'>⚠️ Эти паллеты ожидают обработки лаком перед отправкой на склад</div>" if pallets else ""}
    <div class="card">
      {"<div style='overflow-x:auto'><table><thead><tr><th>ID</th><th>Дата</th><th>Смена</th><th>Декор</th><th>Паллета</th><th>Листов</th><th>Брак</th><th>Оператор</th><th>Статус</th></tr></thead><tbody>"+rows+"</tbody></table></div>" if rows else "<p style='color:var(--mut);text-align:center;padding:20px'>✅ Транзитная зона пуста — все паллеты обработаны</p>"}
    </div>"""
    return render("/transit", content)


# ── СКЛАД ─────────────────────────────────────────────────────────────────────

@app.route("/warehouse", methods=["GET","POST"])
@require_auth
def warehouse_page():
    decor_filter = request.args.get("decor","")
    pallets = get_warehouse_pallets(decor_filter if decor_filter else None)
    summary = get_warehouse_summary()
    total = get_warehouse_total()

    sum_rows = "".join(f'<tr><td><b>{s["decor"]}</b></td><td>{s["total_qty"]}</td>'
                       f'<td>{s["pallets"]}</td><td style="color:var(--mut)">{s["last_date"][:10] if s["last_date"] else "—"}</td></tr>'
                       for s in summary)

    pal_rows = ""
    for p in pallets:
        pal_rows += f"""<tr>
          <td>{p["date"]}</td>
          <td><b>{p["decor"]}</b></td>
          <td>#{p["pallet_num"]}</td>
          <td><b>{p["qty"]}</b></td>
          <td style="color:var(--mut)">{p["lac_operator"] or "—"}</td>
          <td style="color:var(--mut);font-size:11px">{(p["lacquered_at"] or "")[:16]}</td>
          <td><span class="badge warehouse-badge">На складе</span></td>
        </tr>"""

    decors = list(set(s["decor"] for s in summary))
    options = '<option value="">Все декоры</option>' + "".join(
        f'<option value="{d}" {"selected" if d==decor_filter else ""}>{d}</option>' for d in decors)

    # Handle manual add
    add_msg = ""
    if request.method == "POST" and request.form.get("action") == "add_manual":
        try:
            from datetime import date as dt
            save_pallet(
                request.form.get("date", dt.today().strftime("%d.%m.%Y")),
                request.form.get("shift", "день"),
                request.form.get("operator", "Ручной ввод"),
                request.form["decor"],
                int(request.form.get("length", 1220)),
                float(request.form.get("thickness", 2)),
                float(request.form.get("overlay", 0.25)),
                int(request.form.get("pallet_num", 1)),
                int(request.form["qty"]),
                int(request.form.get("defect", 0)),
                request.form.get("time_str", ""),
            )
            # Move directly to warehouse
            from db.database import get_conn
            conn = get_conn(); c = conn.cursor()
            from db.database import P
            c.execute(f"UPDATE extrusion_pallets SET status='warehouse' WHERE decor={P} AND status='transit' AND operator='Ручной ввод' ORDER BY id DESC LIMIT 1",
                      (request.form["decor"],))
            from db.database import get_conn as gc2
            conn.commit(); conn.close()
            add_msg = '<div style="color:var(--grn);margin-bottom:12px;padding:10px;background:var(--grn-bg);border-radius:8px">✅ Паллета добавлена на склад</div>'
        except Exception as e:
            add_msg = f'<div style="color:var(--red);margin-bottom:12px;padding:10px;background:var(--red-bg);border-radius:8px">Ошибка: {e}</div>'

    # Manual add form
    manual_form = """
    <div class="card" style="margin-bottom:12px">
      <h2>➕ Добавить паллету вручную</h2>
      <form method="post" style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;align-items:end">
        <input type="hidden" name="action" value="add_manual">
        <div class="fr"><label>Декор</label><input name="decor" placeholder="82019-9" required></div>
        <div class="fr"><label>Кол-во листов</label><input name="qty" type="number" placeholder="150" required></div>
        <div class="fr"><label>Брак</label><input name="defect" type="number" value="0"></div>
        <div class="fr"><label>Дата</label><input name="date" placeholder="22.04.2026"></div>
        <div class="fr"><label>Оператор</label><input name="operator" value="Ручной ввод"></div>
        <div class="fr"><label>&nbsp;</label><button type="submit" class="btn btn-grn" style="width:100%">Добавить на склад</button></div>
      </form>
    </div>"""

    content = f"""
    {add_msg}
    {manual_form}
    <div class="g4">
      <div class="met grn"><div class="lb">Листов на складе</div><div class="vl">{total["total"] or 0}</div><div class="sb">всего</div></div>
      <div class="met"><div class="lb">Паллет</div><div class="vl">{total["pallets"] or 0}</div></div>
      <div class="met"><div class="lb">Декоров</div><div class="vl">{len(summary)}</div></div>
    </div>
    <div class="two">
      <div class="card"><h2>Сводка по декорам</h2>
        {"<table><thead><tr><th>Декор</th><th>Листов</th><th>Паллет</th><th>Последнее</th></tr></thead><tbody>"+sum_rows+"</tbody></table>" if sum_rows else "<p style='color:var(--mut)'>Склад пуст</p>"}
      </div>
      <div class="card">
        <h2>Паллеты на складе</h2>
        <form method="get" style="display:flex;gap:6px;margin-bottom:10px">
          <select name="decor" style="flex:1">{options}</select>
          <button type="submit" class="btn btn-p">Фильтр</button>
          <a href="/warehouse" class="btn">Все</a>
        </form>
        {"<div style='overflow-x:auto'><table><thead><tr><th>Дата</th><th>Декор</th><th>Паллета</th><th>Листов</th><th>Оператор лака</th><th>Обработано</th><th>Статус</th></tr></thead><tbody>"+pal_rows+"</tbody></table></div>" if pal_rows else "<p style='color:var(--mut)'>Нет данных</p>"}
      </div>
    </div>"""
    return render("/warehouse", content)


# ── ЛАК ───────────────────────────────────────────────────────────────────────

@app.route("/lacquer")
@require_auth
def lacquer_page():
    records = get_lacquer_records(100)
    rows = "".join(f'<tr><td>{r["processed_at"][:16]}</td><td><b>{r["decor"]}</b></td>'
                   f'<td>#{r["pallet_num"]}</td><td>{r["qty"]}</td><td>{r["operator"]}</td></tr>'
                   for r in records)
    content = f"""
    <div class="card"><h2>История обработки лаком ({len(records)} записей)</h2>
      {"<table><thead><tr><th>Время</th><th>Декор</th><th>Паллета</th><th>Листов</th><th>Оператор</th></tr></thead><tbody>"+rows+"</tbody></table>" if rows else "<p style='color:var(--mut);text-align:center;padding:20px'>Нет записей</p>"}
    </div>"""
    return render("/lacquer", content)


# ── ЭКСПОРТ В EXCEL (через ИИ-ассистент) ─────────────────────────────────────

@app.route("/export", methods=["GET","POST"])
@require_auth
def export_page():
    ai_response = ""
    if request.method == "POST" and request.form.get("action") == "ai":
        question = request.form.get("question","")
        ai_response = get_ai_answer(question)

    content = f"""
    <div class="two">
      <div class="card">
        <h2>📥 Выгрузка в Excel</h2>
        <p style="font-size:13px;color:var(--mut);margin-bottom:14px">Скачайте данные за выбранный период</p>
        <form method="get" action="/export/download" style="display:flex;flex-direction:column;gap:10px">
          <div class="fr"><label>Дата начала</label><input type="text" name="date_from" placeholder="дд.мм.гггг"></div>
          <div class="fr"><label>Дата конца</label><input type="text" name="date_to" placeholder="дд.мм.гггг"></div>
          <div class="fr"><label>Что выгружать</label>
            <select name="report_type">
              <option value="all">Всё (склад + выпуск + замесы)</option>
              <option value="warehouse">Только склад</option>
              <option value="production">Только выпуск по участкам</option>
              <option value="mixer">Только замесы</option>
            </select>
          </div>
          <button type="submit" class="btn btn-grn">📥 Скачать Excel</button>
        </form>
      </div>
      <div class="card">
        <h2>🤖 ИИ-ассистент</h2>
        <p style="font-size:13px;color:var(--mut);margin-bottom:12px">Задайте вопрос по данным производства</p>
        <form method="post">
          <input type="hidden" name="action" value="ai">
          <div class="fr">
            <textarea name="question" rows="3" placeholder="Например: Сколько листов выпущено за эту неделю? Какой декор занимает больше всего места на складе?">{request.form.get("question","")}</textarea>
          </div>
          <button type="submit" class="btn btn-p">Спросить</button>
        </form>
        {"<div class='ai-box'><div style='font-size:11px;color:var(--mut);margin-bottom:6px'>Ответ ИИ:</div><div class='ai-resp'>"+ai_response+"</div></div>" if ai_response else ""}
      </div>
    </div>"""
    return render("/export", content)


@app.route("/export/download")
@require_auth
def export_download():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return '<h3 style="font-family:sans-serif;padding:40px;color:#A32D2D">❌ Модуль openpyxl не установлен.<br><br>Обновите requirements.txt и задеплойте заново.</h3>', 500

    report_type = request.args.get("report_type","all")
    date_from = request.args.get("date_from","")
    date_to = request.args.get("date_to","")

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="185FA5")

    def add_sheet(ws, headers, rows):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or ""))+4, 12)

    data = get_full_export(date_from if date_from else None)

    if report_type in ("all","warehouse","production"):
        ws1 = wb.active; ws1.title = "Склад и выпуск"
        headers = ["Дата","Смена","Декор","Длина","Толщ.","Оверлайн","Паллета","Листов","Брак","Статус","Оператор экстр.","Оператор лака","Дата лака"]
        rows = [[p.get("date"),p.get("shift"),p.get("decor"),p.get("length"),p.get("thickness"),
                 p.get("overlay"),p.get("pallet_num"),p.get("qty"),p.get("defect"),
                 p.get("status"),p.get("operator"),p.get("lac_operator"),
                 (p.get("lac_date") or "")[:16]] for p in data["pallets"]]
        add_sheet(ws1, headers, rows)

    if report_type in ("all","mixer"):
        ws2 = wb.create_sheet("Замесы")
        headers2 = ["Дата","Смена","Оператор","Замесов"]
        rows2 = [[m.get("date"),m.get("shift"),m.get("operator"),m.get("batches")] for m in data["mixer"]]
        add_sheet(ws2, headers2, rows2)

    if report_type in ("all","warehouse"):
        ws3 = wb.create_sheet("Сводка склад")
        summary = get_warehouse_summary()
        add_sheet(ws3, ["Декор","Листов","Паллет"],
                  [[s["decor"],s["total_qty"],s["pallets"]] for s in summary])

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"LVT_export_{date.today().strftime('%d%m%Y')}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=fname)


def get_ai_answer(question):
    """Отвечает на вопросы по данным производства"""
    try:
        import urllib.request, json as js
        data = get_full_export()
        warehouse = get_warehouse_summary()
        total = get_warehouse_total()
        transit = get_transit_pallets()

        context = f"""
Данные производства ЛВТ-ламината:
- На складе: {total['total']} листов, {total['pallets']} паллет
- В транзите: {len(transit)} паллет
- Склад по декорам: {js.dumps([{"декор":s["decor"],"листов":s["total_qty"],"паллет":s["pallets"]} for s in warehouse], ensure_ascii=False)}
- Последние паллеты: {js.dumps([{"дата":p["date"],"декор":p["decor"],"листов":p["qty"],"статус":p["status"]} for p in data["pallets"][:20]], ensure_ascii=False)}
- Последние замесы: {js.dumps([{"дата":m["date"],"смена":m["shift"],"замесов":m["batches"]} for m in data["mixer"][:10]], ensure_ascii=False)}
"""
        payload = js.dumps({
            "model": "claude-sonnet-4-20250514",
            "max_tokens": 1000,
            "messages": [{"role":"user","content":f"{context}\n\nВопрос: {question}"}]
        }).encode()
        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=payload,
            headers={"Content-Type":"application/json"},
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = js.loads(resp.read())
            return result["content"][0]["text"]
    except Exception as e:
        return f"ИИ-ассистент временно недоступен: {e}"


# ── ОСТАЛЬНЫЕ РАЗДЕЛЫ ─────────────────────────────────────────────────────────

@app.route("/week")
@require_auth
def week_page():
    ws = request.args.get("week", date.today().strftime("%d.%m.%Y"))
    tasks = get_tasks(ws)
    prog = ""
    for t in tasks:
        done = sum(p["qty"] for p in get_pallets() if p["decor"]==t["decor"])
        pct = min(round(done/t["plan_qty"]*100) if t["plan_qty"] else 0, 100)
        col = prog_color(pct)
        b = f'<span class="badge {"ok" if pct>=80 else ("warn" if pct>=50 else "err")}">{pct}%</span>'
        prog += f'<div class="prog-wrap"><div class="prog-lbl"><span><b>{t["decor"]}</b> · {t["length"]}мм</span><span>{done}/{t["plan_qty"]} {b}</span></div><div class="prog-bar"><div class="prog-fill" style="width:{pct}%;background:{col}"></div></div></div>'
    content = f"""
    <div class="dbar"><span>Неделя с:</span>
      <form method="get" style="display:flex;gap:6px">
        <input name="week" value="{ws}" style="width:120px">
        <button type="submit" class="btn btn-p">Показать</button>
      </form>
    </div>
    <div class="card"><h2>Выполнение задания · {ws}</h2>
      {prog if prog else "<p style='color:var(--mut);text-align:center;padding:20px'>Задание не задано</p>"}
    </div>"""
    return render("/week", content)

@app.route("/mixer")
@require_auth
def mixer_page():
    shifts = get_mixer_shifts()
    rows = "".join(f'<tr><td>{s["date"]}</td>'
                   f'<td><span class="badge {"day" if s["shift"]=="день" else "night"}">{s["shift"].capitalize()}</span></td>'
                   f'<td>{s["operator"]}</td><td><b>{s["batches"]}</b></td>'
                   f'<td style="color:var(--mut)">{round(s["batches"]*434.06):.0f} кг</td></tr>'
                   for s in shifts)
    content = f'<div class="card"><h2>История замесов</h2>{"<table><thead><tr><th>Дата</th><th>Смена</th><th>Оператор</th><th>Замесов</th><th>Сырьё</th></tr></thead><tbody>"+rows+"</tbody></table>" if rows else "<p style=color:var(--mut)>Нет данных</p>"}</div>'
    return render("/mixer", content)

@app.route("/extruder")
@require_auth
def extruder_page():
    df = request.args.get("date","")
    pallets = get_pallets(df if df else None)
    rows = "".join(f'<tr><td>{p["date"]}</td>'
                   f'<td><span class="badge {"day" if p["shift"]=="день" else "night"}">{p["shift"].capitalize()}</span></td>'
                   f'<td><b>{p["decor"]}</b></td><td>#{p["pallet_num"]}</td>'
                   f'<td><b>{p["qty"]}</b></td>'
                   f'<td>{"<span class=badge err>"+str(p["defect"])+"</span>" if p["defect"] else "0"}</td>'
                   f'<td><span class="badge {"transit-badge" if p["status"]=="transit" else "warehouse-badge"}">{p["status"]}</span></td>'
                   f'<td style="color:var(--mut)">{p["operator"]}</td></tr>'
                   for p in pallets)
    content = f"""
    <div class="dbar">
      <form method="get" style="display:flex;gap:6px">
        <input name="date" value="{df}" placeholder="дд.мм.гггг" style="width:120px">
        <button type="submit" class="btn btn-p">Фильтр</button>
        <a href="/extruder" class="btn">Все</a>
      </form>
    </div>
    <div class="card"><h2>Паллеты экструзии</h2>
      {"<div style='overflow-x:auto'><table><thead><tr><th>Дата</th><th>Смена</th><th>Декор</th><th>Паллета</th><th>Листов</th><th>Брак</th><th>Статус</th><th>Оператор</th></tr></thead><tbody>"+rows+"</tbody></table></div>" if rows else "<p style='color:var(--mut)'>Нет данных</p>"}
    </div>"""
    return render("/extruder", content)

@app.route("/tasks", methods=["GET","POST"])
@require_auth
def tasks_page():
    msg = ""
    if request.method == "POST":
        try:
            save_task(request.form["week_start"],request.form["decor"],
                      int(request.form["length"]),float(request.form["thickness"]),
                      float(request.form["overlay"]),int(request.form["plan_qty"]))
            msg = '<div style="color:var(--grn);margin-bottom:12px">✅ Позиция добавлена</div>'
        except Exception as e:
            msg = f'<div style="color:var(--red);margin-bottom:12px">Ошибка: {e}</div>'
    ws = request.args.get("week","")
    tasks = get_tasks(ws) if ws else []
    task_rows = "".join(f'<tr><td>{t["decor"]}</td><td>{t["length"]}</td><td>{t["thickness"]}</td><td>{t["overlay"]}</td><td><b>{t["plan_qty"]}</b></td></tr>' for t in tasks)
    content = f"""{msg}
    <div class="two">
      <div class="card"><h2>Добавить позицию</h2>
        <form method="post">
          <div class="fr"><label>Неделя с (дд.мм.гггг)</label><input name="week_start" value="{ws}" required></div>
          <div class="fr"><label>Декор</label><input name="decor" placeholder="82019-9" required></div>
          <div class="fr"><label>Длина (мм)</label><input name="length" value="1220" required></div>
          <div class="fr"><label>Толщина</label><input name="thickness" value="2" required></div>
          <div class="fr"><label>Оверлайн</label><input name="overlay" value="0.25" required></div>
          <div class="fr"><label>Плановое кол-во</label><input name="plan_qty" required></div>
          <button type="submit" class="btn btn-p">Добавить</button>
        </form>
      </div>
      <div class="card"><h2>Задание</h2>
        <form method="get" style="display:flex;gap:6px;margin-bottom:10px">
          <input name="week" value="{ws}" placeholder="дд.мм.гггг">
          <button type="submit" class="btn">Загрузить</button>
        </form>
        {"<table><thead><tr><th>Декор</th><th>Длина</th><th>Толщ.</th><th>Овер.</th><th>План</th></tr></thead><tbody>"+task_rows+"</tbody></table>" if task_rows else "<p style='color:var(--mut)'>Не задано</p>"}
      </div>
    </div>"""
    return render("/tasks", content)


@app.route("/users", methods=["GET","POST"])
@require_auth
def users_page():
    msg = ""
    if request.method == "POST":
        try:
            tg_id = int(request.form.get("tg_id", 0))
            if tg_id:
                revoke_user(tg_id)
                msg = f'<div style="color:var(--grn);margin-bottom:12px">✅ Пользователь {tg_id} удалён</div>'
        except Exception as e:
            msg = f'<div style="color:var(--red);margin-bottom:12px">Ошибка: {e}</div>'

    users = get_all_users()
    codes = get_codes()

    ROLE_NAMES = {"mixer":"🧪 Миксерщик","extruder":"⚙️ Экструзия","lacquer":"🎨 Лак","boss":"👑 Руководитель"}

    user_rows = ""
    for u in users:
        role_name = ROLE_NAMES.get(u["role"], u["role"])
        username = f"@{u['username']}" if u.get("username") else "—"
        auth_date = str(u.get("authorized_at",""))[:16]
        user_rows += f"""<tr>
            <td><b>{u["name"]}</b><br><span style="color:var(--mut);font-size:11px">{username}</span></td>
            <td>{role_name}</td>
            <td style="font-size:11px;color:var(--mut)">{auth_date}</td>
            <td><code style="font-size:11px">{u["tg_id"]}</code></td>
            <td>
                <form method="post" onsubmit="return confirm('Удалить {u["name"]}?')" style="display:inline">
                    <input type="hidden" name="tg_id" value="{u["tg_id"]}">
                    <button type="submit" class="btn" style="color:var(--red);border-color:var(--red);padding:3px 10px;font-size:11px">
                        🗑 Удалить
                    </button>
                </form>
            </td>
        </tr>"""

    code_rows = ""
    for c in codes:
        role_name = ROLE_NAMES.get(c["role"], c["role"])
        code_rows += f"""<tr>
            <td>{role_name}</td>
            <td><code style="background:var(--bg);padding:3px 8px;border-radius:4px;font-size:13px">{c["code"]}</code></td>
        </tr>"""

    content_html = f"""{msg}
    <div class="two">
        <div class="card">
            <h2>👥 Авторизованные пользователи ({len(users)})</h2>
            {"<div style='overflow-x:auto'><table><thead><tr><th>Имя</th><th>Роль</th><th>Вошёл</th><th>Telegram ID</th><th>Действие</th></tr></thead><tbody>" + user_rows + "</tbody></table></div>" if user_rows else "<p style='color:var(--mut);text-align:center;padding:20px'>Нет пользователей</p>"}
        </div>
        <div class="card">
            <h2>🔑 Коды доступа</h2>
            <p style="font-size:12px;color:var(--mut);margin-bottom:10px">Раздайте эти коды сотрудникам. Код вводится один раз при /start</p>
            <table><thead><tr><th>Роль</th><th>Код</th></tr></thead>
            <tbody>{code_rows}</tbody></table>
            <div style="margin-top:14px;padding-top:12px;border-top:1px solid var(--bor);font-size:12px;color:var(--mut)">
                Сменить код в боте: <code>/code mixer новый_код</code>
            </div>
        </div>
    </div>"""
    return render("/users", content_html)

@app.route("/recipe", methods=["GET","POST"])
@require_auth
def recipe_page():
    msg = ""
    if request.method == "POST":
        try:
            update_recipe(request.form["component"], float(request.form["kg"]))
            msg = '<div style="color:var(--grn);margin-bottom:12px">✅ Обновлено</div>'
        except Exception as e:
            msg = f'<div style="color:var(--red);margin-bottom:12px">Ошибка: {e}</div>'
    recipe = get_recipe(); total = sum(r["kg_per_batch"] for r in recipe)
    rows = "".join(f'<tr><td>{r["component"]}</td><td><b>{r["kg_per_batch"]}</b> кг</td><td style="color:var(--mut)">{r["updated_at"][:16]}</td></tr>' for r in recipe)
    options = "".join(f'<option value="{r["component"]}">{r["component"]}</option>' for r in recipe)
    content = f"""{msg}
    <div class="two">
      <div class="card"><h2>Рецептура (на 1 замес)</h2>
        <table><thead><tr><th>Компонент</th><th>кг/замес</th><th>Обновлено</th></tr></thead>
        <tbody>{rows}<tr style="background:var(--bg)"><td><b>Итого</b></td><td><b style="color:var(--acc)">{total:.2f} кг</b></td><td></td></tr></tbody></table>
      </div>
      <div class="card"><h2>Изменить компонент</h2>
        <form method="post">
          <div class="fr"><label>Компонент</label><select name="component">{options}</select></div>
          <div class="fr"><label>Новое значение (кг/замес)</label><input name="kg" type="number" step="0.01" required></div>
          <button type="submit" class="btn btn-p">Сохранить</button>
        </form>
      </div>
    </div>"""
    return render("/recipe", content)


# ── API для бота ──────────────────────────────────────────────────────────────

@app.route("/api/tasks")
@require_auth
def api_tasks():
    """Возвращает актуальное задание в JSON для бота"""
    tasks = get_tasks()  # берёт последнее задание
    result = []
    for t in tasks:
        # Считаем прогресс
        pallets = get_pallets()
        done_qty = sum(p["qty"] for p in pallets if p["decor"] == t["decor"])
        done_cnt = sum(1 for p in pallets if p["decor"] == t["decor"])
        plan = t["plan_qty"]
        pallets_needed = -(-plan // 150)
        result.append({
            "id": t["id"],
            "week_start": t["week_start"],
            "decor": t["decor"],
            "length": t["length"],
            "thickness": t["thickness"],
            "overlay": t["overlay"],
            "plan_qty": plan,
            "done_qty": done_qty,
            "done_cnt": done_cnt,
            "pallets_needed": pallets_needed,
            "pct": round(done_qty / plan * 100) if plan else 0,
            "done": done_qty >= plan,
        })
    from flask import jsonify
    return jsonify({"tasks": result, "week": result[0]["week_start"] if result else ""})

@app.route("/api/complete_pallet", methods=["POST"])
@require_auth
def api_complete_pallet():
    """Оператор отметил паллету как выполненную"""
    from flask import jsonify
    data = request.get_json()
    task_id = data.get("task_id")
    operator = data.get("operator", "")
    qty = data.get("qty", 150)
    result = complete_task_pallet(task_id, operator, qty)
    if result:
        return jsonify({"ok": True, "result": result})
    return jsonify({"ok": False, "error": "Task not found"}), 404

if __name__ == "__main__":
    init_db()
    port = int(os.getenv("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
